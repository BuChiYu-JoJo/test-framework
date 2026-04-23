# -*- coding: utf-8 -*-
"""
Excel Parser - Excel格式用例解析器
支持 L1 零门槛 Excel 用例格式

Excel 格式（设计文档 v1.0 L1）：
| 用例ID | 用例名称 | 模块 | 优先级 | 操作类型 | 目标元素 | 输入值 | 预期结果 | 备注 |
|--------|---------|------|--------|---------|---------|--------|---------|-----|

同一用例ID的多行组成完整用例（setup → steps → teardown）
"""

import re
from pathlib import Path
from typing import Any, Dict, List
from datetime import datetime


class ExcelParser:
    """Excel 用例解析器（L1 格式）"""

    # Excel 列索引（0-based）
    COL_ID = 0           # 用例ID
    COL_NAME = 1         # 用例名称
    COL_MODULE = 2       # 模块
    COL_PRIORITY = 3    # 优先级
    COL_ACTION = 4      # 操作类型
    COL_TARGET = 5      # 目标元素
    COL_VALUE = 6       # 输入值
    COL_EXPECTED = 7    # 预期结果
    COL_REMARK = 8       # 备注

    # 操作类型映射（中文 → 英文关键字）
    ACTION_MAP = {
        # 导航
        '导航': 'navigate',
        '打开': 'navigate',
        'goto': 'navigate',
        # Setup/Teardown
        '清除cookie': 'clear_browser_cookies',
        'clear_browser_cookies': 'clear_browser_cookies',
        '清除cookies': 'clear_browser_cookies',
        '清除缓存': 'clear_browser_cookies',
        '登出': 'logout',
        # 点击
        '点击': 'click',
        'click': 'click',
        '双击': 'dblclick',
        'dblclick': 'dblclick',
        '右键': 'rightclick',
        'rightclick': 'rightclick',
        # 输入
        '输入': 'type',
        'type': 'type',
        '清空': 'clear',
        'clear': 'clear',
        # 选择
        '选择': 'select',
        'select': 'select',
        # 勾选
        '勾选': 'check',
        'check': 'check',
        '取消勾选': 'uncheck',
        'uncheck': 'uncheck',
        # 等待
        '等待': 'wait',
        'wait': 'wait',
        '等待元素': 'wait_for',
        'wait_for': 'wait_for',
        '等待URL': 'wait_for_url',
        'wait_for_url': 'wait_for_url',
        # 断言
        '断言文本': 'assert_text',
        'assert_text': 'assert_text',
        '断言可见': 'assert_visible',
        'assert_visible': 'assert_visible',
        '断言隐藏': 'assert_hidden',
        'assert_hidden': 'assert_hidden',
        '断言URL': 'assert_url',
        'assert_url': 'assert_url',
        '断言数量': 'assert_count',
        'assert_count': 'assert_count',
        # API
        'api请求': 'api_request',
        'api_request': 'api_request',
        # 执行
        '执行js': 'execute_js',
        'execute_js': 'execute_js',
        # 截图
        '截图': 'screenshot',
        'screenshot': 'screenshot',
        # 上传
        '上传': 'upload',
        'upload': 'upload',
        # 滚动
        '滚动': 'scroll',
        'scroll': 'scroll',
        # 悬停
        '悬停': 'hover',
        'hover': 'hover',
        # 刷新
        '刷新': 'refresh',
        'refresh': 'refresh',
        # 后退
        '后退': 'back',
        'back': 'back',
        # 前进
        '前进': 'forward',
        'forward': 'forward',
        # 取消选择
        '取消选择': 'deselect',
        'deselect': 'deselect',
    }

    def parse(self, file_path: str) -> List[Dict]:
        """
        解析 Excel 用例文件

        Args:
            file_path: Excel 文件路径（.xlsx / .xls）

        Returns:
            用例数据字典列表（每个用例一个 dict）
            格式与 YamlParser.parse() 输出一致
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        suffix = path.suffix.lower()
        if suffix == '.xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
        elif suffix == '.xls':
            import xlrd
            rb = xlrd.open_workbook(path)
            wb = self._xls_to_xlsx(rb)
        else:
            raise ValueError(f"Unsupported Excel format: {suffix}")

        # 读取第一个 sheet
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            raise ValueError("Empty Excel file")

        # 解析用例
        return self._parse_rows(rows)

    def _xls_to_xlsx(self, rb) -> 'openpyxl.Workbook':
        """将旧版 .xls 转为 openpyxl.Workbook"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active

        for sheet_idx in range(rb.nsheets):
            sheet = rb.sheet_by_index(sheet_idx)
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                ws.append(row)

            if sheet_idx < rb.nsheets - 1:
                wb.create_sheet()

        return wb

    def _parse_rows(self, rows: List[tuple]) -> List[Dict]:
        """
        将行数据按用例ID分组，转换为标准化用例格式

        Returns:
            用例列表
        """
        # 跳过表头（第1行）
        if self._is_header_row(rows[0]):
            data_rows = rows[1:]
        else:
            data_rows = rows

        # 按用例ID分组
        case_groups: Dict[str, List[Dict]] = {}
        for row in data_rows:
            if not row or all(v is None for v in row):
                continue  # 跳过空行

            step = self._parse_row(row)
            if not step:
                continue

            case_id = step.get('case_id', '')
            if not case_id:
                continue

            if case_id not in case_groups:
                case_groups[case_id] = []
            case_groups[case_id].append(step)

        # 转换为标准化格式
        cases = []
        for case_id, steps in case_groups.items():
            case = self._build_case(case_id, steps)
            cases.append(case)

        return cases

    def _is_header_row(self, row: tuple) -> bool:
        """判断是否为表头行"""
        if not row or not row[0]:
            return False
        first_cell = str(row[0]).strip().lower()
        return first_cell in ('用例id', '用例名称', 'case_id', 'casename', 'id')

    def _parse_row(self, row: tuple) -> Dict:
        """解析单行数据为步骤字典"""
        if len(row) <= self.COL_ID:
            return {}

        case_id = str(row[self.COL_ID]).strip() if row[self.COL_ID] is not None else ''
        if not case_id:
            return {}

        # 解析操作类型
        action_raw = str(row[self.COL_ACTION]).strip() if row[self.COL_ACTION] is not None else ''
        action = self._normalize_action(action_raw)

        # 解析目标元素
        target = str(row[self.COL_TARGET]).strip() if row[self.COL_TARGET] is not None else ''

        # 解析输入值
        value = str(row[self.COL_VALUE]).strip() if row[self.COL_VALUE] is not None else ''

        # 解析预期结果
        expected = str(row[self.COL_EXPECTED]).strip() if row[self.COL_EXPECTED] is not None else ''

        # 解析备注（用于提取 step 类型：setup/step/teardown）
        remark = str(row[self.COL_REMARK]).strip() if row[self.COL_REMARK] is not None else ''

        # 从备注中判断步骤类型
        step_type = self._detect_step_type(remark, action)

        return {
            'case_id': case_id,
            'name': str(row[self.COL_NAME]).strip() if row[self.COL_NAME] is not None else '',
            'module': str(row[self.COL_MODULE]).strip() if row[self.COL_MODULE] is not None else '',
            'priority': str(row[self.COL_PRIORITY]).strip() if row[self.COL_PRIORITY] is not None else 'P2',
            'action': action,
            'target': target,
            'value': value,
            'expected': expected,
            'remark': remark,
            'step_type': step_type,  # setup / step / teardown
        }

    def _normalize_action(self, action_raw: str) -> str:
        """将操作类型映射为英文关键字"""
        if not action_raw:
            return ''

        action_lower = action_raw.strip().lower()

        # 精确匹配
        if action_lower in self.ACTION_MAP:
            return self.ACTION_MAP[action_lower]

        # 模糊匹配（包含关键词）
        for key, value in self.ACTION_MAP.items():
            if key in action_lower or action_lower in key:
                return value

        # 大小写不敏感精确匹配
        for key, value in self.ACTION_MAP.items():
            if key.lower() == action_lower:
                return value

        # 无法识别，返回原始值（小写）
        return action_lower

    def _detect_step_type(self, remark: str, action: str) -> str:
        """
        从备注判断步骤类型

        支持格式：
        - 备注为 "[setup]" 或 "setup" → setup
        - 备注为 "[teardown]" 或 "teardown" → teardown
        - 默认 → step
        """
        if not remark:
            return 'step'

        remark_lower = remark.lower().strip()

        # 显式标记
        if '[setup]' in remark_lower or remark_lower == 'setup':
            return 'setup'
        if '[teardown]' in remark_lower or remark_lower == 'teardown':
            return 'teardown'
        if '[前置]' in remark or '前置' in remark:
            return 'setup'
        if '[后置]' in remark or '后置' in remark:
            return 'teardown'

        return 'step'

    def _build_case(self, case_id: str, steps: List[Dict]) -> Dict:
        """
        将同一用例的多行步骤构建为标准化用例格式

        与 YamlParser 输出一致：
        {
            'id': 'TC001',
            'name': '登录成功',
            'module': 'login',
            'priority': 'P0',
            'tags': [],
            'version': '1.0.0',
            'author': '',
            'created_at': '',
            'setup': [...],
            'steps': [...],
            'teardown': [...],
            'expected': [...],
            'data': {},
            '_raw': {}
        }
        """
        # 从第一行获取用例基础信息
        first = steps[0]
        case_name = first.get('name', case_id)
        module = first.get('module', '')
        priority = first.get('priority', 'P2')

        # 分离 setup / steps / teardown
        setup_steps = []
        main_steps = []
        teardown_steps = []

        for step in steps:
            step_type = step.get('step_type', 'step')
            normalized = self._normalize_step(step)

            if step_type == 'setup':
                setup_steps.append(normalized)
            elif step_type == 'teardown':
                teardown_steps.append(normalized)
            else:
                main_steps.append(normalized)

        # 解析预期结果
        expected_list = []
        for step in steps:
            exp = step.get('expected', '').strip()
            if exp:
                expected_list.append(self._parse_expected(exp))

        # 构建标准化输出
        case = {
            'id': case_id,
            'name': case_name,
            'module': module,
            'priority': priority,
            'tags': [],
            'version': '1.0.0',
            'author': '',
            'created_at': datetime.now().strftime('%Y-%m-%d'),

            'setup': setup_steps,
            'steps': main_steps,
            'teardown': teardown_steps,

            'expected': expected_list,
            'data': {},

            # 备用原始数据
            '_raw': {
                'source': 'excel',
                'case_id': case_id,
                'row_count': len(steps),
            }
        }

        return case

    def _normalize_step(self, step: Dict) -> Dict:
        """
        标准化单个步骤

        Returns:
            {no, action, target, value, description}
        """
        return {
            'no': 0,  # 由引擎填充
            'action': step.get('action', ''),
            'target': step.get('target', ''),
            'value': step.get('value', ''),
            'description': step.get('remark', ''),
        }

    def _parse_expected(self, expected_str: str) -> Dict:
        """
        解析预期结果字符串

        支持格式：
        - "URL包含 /dashboard/" → {url_contains: /dashboard/}
        - "元素可见 .sidebar" → {element_visible: .sidebar}
        - "文本包含 欢迎回来" → {contains: 欢迎回来}
        - "状态 = 成功" → {status: 成功}
        """
        expected_str = expected_str.strip()
        if not expected_str:
            return {}

        # URL相关
        if 'url' in expected_str.lower() or 'URL' in expected_str:
            if '包含' in expected_str:
                parts = expected_str.split('包含')
                if len(parts) > 1:
                    return {'url_contains': parts[-1].strip()}
            if '==' in expected_str:
                parts = expected_str.split('==')
                return {'url_contains': parts[-1].strip()}

        # 元素可见/隐藏
        if '可见' in expected_str or 'visible' in expected_str.lower():
            if '不' in expected_str:
                parts = expected_str.replace('元素不可见', '').replace('not visible', '').strip()
                return {'element_not_visible': parts}
            parts = expected_str.replace('元素可见', '').replace('visible', '').strip()
            return {'element_visible': parts}

        # 文本包含
        if '文本' in expected_str or '包含' in expected_str or 'contains' in expected_str.lower():
            if '包含' in expected_str:
                parts = expected_str.split('包含')
                if len(parts) > 1:
                    return {'contains': parts[-1].strip()}
            # 纯文本
            return {'contains': expected_str}

        # 状态判断
        if '状态' in expected_str or 'status' in expected_str.lower():
            if '=' in expected_str:
                parts = expected_str.split('=')
                return {'status': parts[-1].strip()}
            return {'status': expected_str}

        # 默认：文本断言
        return {'contains': expected_str}

    def parse_case(self, file_path: str, case_id: str) -> Dict:
        """
        解析 Excel 并返回指定用例ID的用例数据

        Args:
            file_path: Excel 文件路径
            case_id: 用例ID

        Returns:
            该用例的标准化数据字典
        """
        cases = self.parse(file_path)
        for case in cases:
            if case.get('id') == case_id:
                return case
        raise ValueError(f"Case not found: {case_id}")


def parse_case_file(file_path: str) -> List[Dict]:
    """
    快捷函数：解析 Excel 用例文件

    Args:
        file_path: Excel 文件路径

    Returns:
        用例数据列表
    """
    parser = ExcelParser()
    return parser.parse(file_path)


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        cases = parse_case_file(sys.argv[1])
        print(f"Found {len(cases)} cases:")
        for case in cases:
            print(f"  [{case['id']}] {case['name']} ({case['module']}) - {case['priority']}")
            print(f"    setup: {len(case['setup'])} steps, steps: {len(case['steps'])} steps, teardown: {len(case['teardown'])} steps")
